from __future__ import annotations

import csv
from collections import Counter
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from normal.movie_plan import parse_movie_name
from normal.quality_review import build_audio_summary
from normal.scan import analyze_library, infer_album_root


MOVIE_STATUS_PRIORITY = {"severe": 0, "review": 1, "ok": 2, "unscored": 3}


def write_collection_csv(source_root: Path, csv_path: Path) -> None:
    report = analyze_library(source_root)
    rows = build_collection_rows(report)
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["album_artist", "album", "date", "genre", "track_count", "path"])
        writer.writerows(rows)


def build_collection_rows(report) -> list[list[str]]:
    rows: list[list[str]] = []
    for album in report.albums:
        album_path = Path(album.path)
        tracks = [track for track in report.tracks if infer_album_root(Path(track.path)) == album_path]
        if not tracks:
            continue
        first_track = tracks[0]
        album_artist = first_track.tags.get("albumartist") or first_track.tags.get("artist", "")
        album_title = first_track.tags.get("album", "")
        release_date = consensus_year_or_date([track.tags.get("date", "") for track in tracks])
        genre = consensus_genre([track.tags.get("genre", "") for track in tracks])
        rows.append(
            [
                album_artist,
                album_title,
                release_date,
                genre,
                str(album.track_count),
                album.path,
            ]
        )

    rows.sort(key=lambda row: (row[0], row[2], row[1], row[5]))
    return rows


def consensus_year_or_date(values: list[str]) -> str:
    normalized = [value.strip() for value in values if value.strip()]
    if not normalized:
        return ""
    counts = Counter(value[:4] if len(value) >= 4 and value[:4].isdigit() else value for value in normalized)
    return sorted(counts.items(), key=lambda item: (-item[1], item[0]))[0][0]


def consensus_genre(values: list[str]) -> str:
    normalized = [value.strip() for value in values if value.strip()]
    if not normalized:
        return ""

    counts = Counter(normalized)
    top_count = max(counts.values())
    leaders = sorted(genre for genre, count in counts.items() if count == top_count)
    if len(leaders) == 1:
        return leaders[0]
    return ";".join(sorted(counts))


def write_movie_review_csv(report_path: Path, csv_path: Path, minimum_status: str = "review") -> None:
    payload = json.loads(report_path.read_text(encoding="utf-8"))
    rows = build_movie_review_rows(payload, minimum_status=minimum_status)
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "status",
                "triage_score",
                "score",
                "replacement_priority_score",
                "replacement_priority_label",
                "replacement_year_hint",
                "confidence",
                "resolution",
                "runtime_minutes",
                "video_kbps",
                "audio_kbps",
                "audio_channels",
                "audio_summary",
                "mb_per_min",
                "container",
                "video_codec",
                "audio_codec",
                "profile_label",
                "weak_candidate",
                "profile_percentile",
                "anchor_distance",
                "playback_risk_count",
                "indexing_visibility_risk_count",
                "standards_review_count",
                "standards_failure_count",
                "diagnostics",
                "domain_results",
                "reasons",
                "path",
            ]
        )
        writer.writerows(rows)


def build_movie_review_rows(payload: dict, minimum_status: str = "review") -> list[list[str]]:
    threshold = MOVIE_STATUS_PRIORITY.get(minimum_status, MOVIE_STATUS_PRIORITY["review"])
    rows: list[list[str]] = []

    for item in payload.get("movies", []):
        review = item.get("review", {})
        status = review.get("status", "unscored")
        profile = item.get("profile", {})
        if MOVIE_STATUS_PRIORITY.get(status, 99) > threshold:
            if not profile:
                continue

        facts = review.get("facts") or item.get("facts", {})
        derived = review.get("derived", {})
        reasons = review.get("reasons", [])
        diagnostics = profile.get("diagnostics", [])
        risk_counts = profile.get("risk_counts", {})
        runtime_seconds = facts.get("runtime_seconds")
        runtime_minutes = ""
        if runtime_seconds:
            runtime_minutes = f"{runtime_seconds / 60:.1f}"

        mb_per_min = derived.get("mb_per_min")
        mb_per_min_display = ""
        if mb_per_min is not None:
            mb_per_min_display = f"{mb_per_min:.1f}"

        rows.append(
            [
                status,
                f"{float(item.get('triage_score', 0.0)):.1f}",
                str(review.get("score", "")),
                f"{float(item.get('replacement_priority_score', 1.0)):.2f}",
                item.get("replacement_priority_label", ""),
                str(item.get("replacement_year_hint", "") or ""),
                review.get("confidence", ""),
                facts.get("resolution_bucket", "") or "",
                runtime_minutes,
                str(facts.get("video_bitrate_kbps", "") or ""),
                str(facts.get("audio_bitrate_kbps", "") or ""),
                str(facts.get("audio_channels", "") or ""),
                facts.get("audio_summary", "") or "",
                mb_per_min_display,
                facts.get("container", "") or "",
                facts.get("video_codec", "") or "",
                facts.get("audio_codec", "") or "",
                profile.get("label", ""),
                str(bool(profile.get("weak_candidate"))).lower(),
                str(profile.get("percentile", "") or ""),
                str(profile.get("anchor_distance", "") or ""),
                str(risk_counts.get("playback_risk", "") or ""),
                str(risk_counts.get("indexing_visibility_risk", "") or ""),
                str(risk_counts.get("standards_review", "") or ""),
                str(risk_counts.get("standards_failure", "") or ""),
                "; ".join(diagnostic.get("code", "") for diagnostic in diagnostics),
                "; ".join(result.get("code", "") for result in (profile.get("domain_results") or []) if isinstance(result, dict)),
                "; ".join(reason.get("code", "") for reason in reasons),
                item.get("path", ""),
            ]
        )

    rows.sort(
        key=lambda row: (
            -safe_float(row[1]),
            MOVIE_STATUS_PRIORITY.get(row[0], 99),
            -safe_int(row[2]),
            row[-1].lower(),
        )
    )
    return rows


def safe_int(value: str) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def safe_float(value: str) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


# ---------------------------------------------------------------------------
# Movie register XLSX
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F2D3D")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_ALT_FILL = PatternFill(fill_type="solid", fgColor="F2F4F7")
_BODY_FONT = Font(name="Calibri", size=11)

_REGISTER_COLUMNS = [
    ("Title", 36),
    ("Year", 7),
    ("Resolution", 12),
    ("Video", 10),
    ("Audio", 16),
    ("Container", 11),
    ("Size", 10),
]


def write_movie_register_xlsx(report_path: Path, xlsx_path: Path) -> None:
    payload = json.loads(report_path.read_text(encoding="utf-8"))
    rows = _build_register_rows(payload)
    wb = Workbook()
    ws = wb.active
    ws.title = "All Movies"

    for col_idx, (header, width) in enumerate(_REGISTER_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_REGISTER_COLUMNS))}1"

    for row_idx, row in enumerate(rows, start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _BODY_FONT
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def _build_register_rows(payload: dict) -> list[list]:
    rows: list[list] = []
    for item in payload.get("movies", []):
        facts = item.get("review", {}).get("facts") or item.get("facts", {})
        title, year = _parse_title_year(item.get("path", ""))
        resolution = facts.get("resolution_bucket") or ""
        video_codec = _display_video_codec(facts.get("video_codec") or "")
        audio = _derive_audio_label(
            facts.get("audio_codec") or "",
            facts.get("audio_channels"),
            facts.get("audio_profile") or "",
            facts.get("audio_summary") or "",
        )
        container = facts.get("container") or ""
        size = _format_file_size(facts.get("file_size_bytes"))
        rows.append([title, year, resolution, video_codec, audio, container, size])

    rows.sort(key=lambda r: (r[0].lower(), r[1]))
    return rows


def _parse_title_year(path: str) -> tuple[str, str]:
    import re
    movie_path = Path(path)
    parsed = parse_movie_name(movie_path)
    if parsed.title is not None and parsed.year is not None:
        return parsed.title, str(parsed.year)
    stem = movie_path.stem
    year_match = re.search(r"\b(19\d{2}|20\d{2}|2100)\b", stem)
    year = year_match.group(1) if year_match else ""
    if year_match:
        title = stem[: year_match.start()].strip(" -_.()")
    else:
        title = stem
    title = re.sub(r"[\._]", " ", title).strip()
    return title, year


def _display_video_codec(codec: str) -> str:
    mapping = {
        "h264": "H.264",
        "hevc": "H.265",
        "av1": "AV1",
        "vp9": "VP9",
        "mpeg4": "MPEG-4",
        "mpeg2video": "MPEG-2",
        "vc1": "VC-1",
        "theora": "Theora",
    }
    return mapping.get(codec.lower(), codec.upper() if codec else "")


def _derive_audio_label(codec: str, channels: int | None, profile: str, summary: str) -> str:
    if summary:
        return summary
    return build_audio_summary(codec, channels, profile)[-1] or ""


def _format_file_size(size_bytes: int | None) -> str:
    if not size_bytes:
        return ""
    gb = size_bytes / (1024 ** 3)
    if gb >= 1:
        return f"{gb:.1f} GB"
    mb = size_bytes / (1024 ** 2)
    return f"{mb:.0f} MB"
